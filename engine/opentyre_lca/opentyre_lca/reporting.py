from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .tyre_model import (
    TyreLCAParams,
    calculate_lca,
    lci_records,
    contribution_records,
    run_scenarios,
)


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LIGHT_FILL = PatternFill("solid", fgColor="F3F6FA")

TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
HEADER_FONT = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
NORMAL_FONT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _title(ws, title: str, subtitle: str | None = None) -> None:
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=11, color="666666")
        ws.merge_cells("A2:H2")


def _write_key_values(ws, start_row: int, rows: list[tuple[str, Any]]) -> int:
    row = start_row
    for key, value in rows:
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)

        ws.cell(row, 1).font = HEADER_FONT
        ws.cell(row, 1).fill = LIGHT_FILL

        for col in [1, 2]:
            c = ws.cell(row, col)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(c.value, float):
                c.number_format = "#,##0.000"

        row += 1

    return row + 1


def _write_table(ws, start_row: int, rows: list[dict]) -> int:
    if not rows:
        ws.cell(start_row, 1, "No data")
        return start_row + 2

    headers = list(rows[0].keys())

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row_data in enumerate(rows, start=start_row + 1):
        for c, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            cell = ws.cell(r, c, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if isinstance(value, float):
                cell.number_format = "#,##0.000"

    ws.freeze_panes = ws.cell(start_row + 1, 1)
    ws.auto_filter.ref = ws.dimensions

    return start_row + len(rows) + 2


def _autosize(ws, max_width: int = 44) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(width, max_width)


def _safe_sheet_name(name: str) -> str:
    invalid = ["\\", "/", "*", "[", "]", ":", "?"]
    for char in invalid:
        name = name.replace(char, "-")
    return name[:31]


def build_customer_workbook(
    params: TyreLCAParams,
    scenario_definitions: dict | None = None,
    customer_name: str = "Customer",
    project_name: str = "End-of-life tyre LCA screening",
    prepared_by: str = "LCA_TYRE / Calvyx",
) -> bytes:
    """
    Build customer-facing Excel workbook report.

    This workbook is for engineering screening and transparent LCI review.
    It is not a certified ISO 14040/14044 LCA report unless independently reviewed.
    """

    result = calculate_lca(params)
    lci = lci_records(params)
    contribution = contribution_records(params, group_by="stage")

    scenarios = []
    if scenario_definitions:
        scenarios = run_scenarios(params, scenario_definitions)

    wb = Workbook()

    # ------------------------------------------------------------------
    # Cover
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Cover"
    _title(ws, "LCA_TYRE Customer Workbook Report", project_name)

    cover_rows = [
        ("Customer", customer_name),
        ("Prepared by", prepared_by),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Model status", "Engineering screening model"),
        ("Certification status", "Not certified ISO 14040/14044 LCA or EPD"),
        ("Important note", "All placeholder emission factors must be verified before external claims."),
    ]
    _write_key_values(ws, 4, cover_rows)
    _autosize(ws)

    # ------------------------------------------------------------------
    # Executive Summary
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Executive Summary")
    _title(ws, "Executive Summary", "Main LCA screening results")

    summary_rows = [
        ("Annual ELT feed, t/y", result["annual_elt_t"]),
        ("Raw rCB production, t/y", result["raw_rcb_tpy"]),
        ("Purified rCB, t/y", result["purified_rcb_tpy"]),
        ("CB product for substitution, t/y", result["cb_product_for_substitution_tpy"]),
        ("Virgin carbon black avoided, t/y", result["virgin_cb_avoided_tpy"]),
        ("Gross burdens, tCO2e/y", result["gross_burdens_tco2e_per_year"]),
        ("Gross credits, tCO2e/y", result["gross_credits_tco2e_per_year"]),
        ("Net impact, tCO2e/y", result["net_impact_tco2e_per_year"]),
        ("Net saving, tCO2e/y", result["net_saving_tco2e_per_year"]),
        ("Net saving, tCO2e/t ELT", result["net_saving_tco2e_per_t_elt"]),
    ]
    _write_key_values(ws, 4, summary_rows)
    _autosize(ws)

    # ------------------------------------------------------------------
    # Inputs
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Inputs")
    _title(ws, "Model Inputs", "Editable assumptions used for this workbook")

    d = params.deashing
    input_rows = [
        ("Case name", params.case_name),
        ("Annual ELT feed, t/y", params.annual_elt_t),
        ("Raw rCB yield, t/t ELT", params.rcb_yield_t_per_t_elt),
        ("Virgin CB EF, tCO2e/t", params.virgin_cb_ef_tco2e_per_t),
        ("Raw rCB pyrolysis EF, tCO2e/t", params.raw_rcb_pyrolysis_ef_tco2e_per_t),
        ("Include shipping credit", params.include_shipping_credit),
        ("Avoided shipping credit, tCO2e/y", params.avoided_shipping_tco2e_per_year),
        ("Include oil credit", params.include_oil_credit),
        ("Oil credit, tCO2e/y", params.oil_credit_tco2e_per_year),
        ("Deashing enabled", d.enabled),
        ("Purified rCB yield, t/t CBp", d.purified_yield_t_per_t_feed),
        ("HNO3, kg/t CBp", d.hno3_kg_per_t_feed),
        ("NaOH, kg/t CBp", d.naoh_kg_per_t_feed),
        ("Electricity, kWh/t CBp", d.electricity_kwh_per_t_feed),
        ("Fresh water, m3/t CBp", d.fresh_water_m3_per_t_feed),
        ("Wastewater, m3/t CBp", d.wastewater_m3_per_t_feed),
        ("Include zinc credit", d.include_zinc_credit),
    ]
    _write_key_values(ws, 4, input_rows)
    _autosize(ws)

    # ------------------------------------------------------------------
    # LCI Inventory
    # ------------------------------------------------------------------
    ws = wb.create_sheet("LCI Inventory")
    _title(ws, "Life-Cycle Inventory", "Burden and credit flows")
    _write_table(ws, 4, lci)
    _autosize(ws)

    # ------------------------------------------------------------------
    # Contribution
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Contribution")
    _title(ws, "Contribution Analysis", "Impact by life-cycle stage")
    _write_table(ws, 4, contribution)
    _autosize(ws)

    # ------------------------------------------------------------------
    # Scenarios
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Scenarios")
    _title(ws, "Scenario Results", "Comparison of model scenarios")
    if scenarios:
        _write_table(ws, 4, scenarios)
    else:
        ws["A4"] = "No scenario definitions supplied."
    _autosize(ws)

    # ------------------------------------------------------------------
    # Assumptions
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Assumptions")
    _title(ws, "Assumptions and Limitations")

    assumptions = [
        ("Functional unit", "Annual treatment of end-of-life tyres, normalised per tonne ELT."),
        ("Impact category", "GWP100, expressed as kg CO2e or t CO2e."),
        ("Model type", "Engineering screening LCI/LCA model."),
        ("Allocation", "Current default uses system expansion for avoided virgin carbon black."),
        ("Data status", "Emission factors are placeholders unless explicitly verified and referenced."),
        ("Certification status", "Not certified ISO 14040/14044 LCA or EPD."),
        ("Review requirement", "External claims require reviewed data, documented system boundary and independent verification."),
    ]
    _write_key_values(ws, 4, assumptions)
    _autosize(ws)

    # ------------------------------------------------------------------
    # References
    # ------------------------------------------------------------------
    ws = wb.create_sheet("References")
    _title(ws, "References and Data Quality")
    reference_rows = [
        {
            "reference_id": "REF-001",
            "item": "Virgin carbon black emission factor",
            "status": "Placeholder",
            "source": "To be verified",
            "notes": "Replace with verified LCI database, EPD, or literature value.",
        },
        {
            "reference_id": "REF-002",
            "item": "Raw rCB pyrolysis emission factor",
            "status": "Placeholder",
            "source": "Engineering model",
            "notes": "Replace with plant-specific energy and material inventory.",
        },
        {
            "reference_id": "REF-003",
            "item": "Deashing chemicals",
            "status": "Placeholder",
            "source": "Engineering workbook defaults",
            "notes": "Replace HNO3, NaOH, electricity, water and wastewater factors with verified values.",
        },
    ]
    _write_table(ws, 4, reference_rows)
    _autosize(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
